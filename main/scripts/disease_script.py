import openpyxl
from CirculoInfantil.settings import BASE_DIR
from main.models.disease import Disease
from main.models.intolerance import Intolerance


def save_disease_from_xl():
    path = '{}{}'.format(BASE_DIR, '/main/scripts/file/disease.xlsx')
    wb = openpyxl.load_workbook(path)
    sheet = wb.active
    for i in range(1, sheet.max_row + 1):
        code_disease = sheet.cell(row=i, column=1).value
        name_disease = sheet.cell(row=i, column=2).value
        description_disease = sheet.cell(row=i, column=3).value
        Disease.objects.get_or_create(code_disease=code_disease, name_disease=name_disease,
                                      description_disease=description_disease)


def save_intolerance_from_xl():
    path = '{}{}'.format(BASE_DIR, '/main/scripts/file/intolerance.xlsx')
    workbook = openpyxl.load_workbook(path)
    sheet = workbook.active
    for i in range(1, sheet.max_row + 1):
        code_into = sheet.cell(row=i, column=1).value
        name_intolerance = sheet.cell(row=i, column=2).value
        description_intolerance = sheet.cell(row=i, column=3).value
        Intolerance.objects.get_or_create(code_into=code_into, name_intolerance=name_intolerance,
                                          description_intolerance=description_intolerance)
